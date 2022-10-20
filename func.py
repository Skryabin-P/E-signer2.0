import os
import time

import pandas as pd
import tempfile
import pdfkit
from win32com import client
from openpyxl import load_workbook
import subprocess
def initial_layout(sg):
    choose_sheet = [[sg.Text('Выберете лист для подписи'),sg.Combo(key='sheet_list',size=(25,3),values=[])]]
    color = 'black'
    common_layout = [
        [sg.Text('Выберете файл для подписания'), sg.Input(key='FILE', enable_events=True, readonly=True,text_color=color),
         sg.FileBrowse('выбрать',key='browse_files', file_types=[('Файлы для подписания', '*.pdf')])],
        [sg.Text('Перед подписанием закройте PDF файл подлежащий подписанию!',key='caution',background_color='red',text_color='white')],
        [sg.Column(choose_sheet,key='sheets',visible=False)],
        [sg.Text('Выберете сертификат для подписания'), sg.Input(key='PFX', readonly=True,text_color=color),
         sg.FileBrowse('выбрать', file_types=[('Сертификат', "*.pfx")])],
        [sg.Text('Введите пароль от сертификата'), sg.InputText(key='PASSWORD', password_char='*')],
        [sg.Text('Выберете папку куда сохранить подписанный документ'), sg.Input(key='FOLDER', readonly=True,text_color=color ),
         sg.FolderBrowse('выбрать')],
        # [sg.ProgressBar(max_value=100)],
        [sg.Checkbox(text = 'Открыть файл после подписания', key='open', default=True)],
        [sg.Button('Подписать', key='sign'), sg.Button('Сбросить', key='clear'), sg.Exit('Выйти', key='exit')],
        [sg.Text('Ожидайте завершения...',key='waiting',background_color='green',text_color='orange',visible=False)],
        [sg.Text('Успешно подписан!',key='success',background_color='green',text_color='orange',visible=False)]]
    return common_layout

def clear_input(window,sg):
    window['sheets'].update(visible=False)
    for key, element in window.key_dict.items():
        if isinstance(element, sg.Input):
            element.update(value='')


def check_input(values,sg):
    keys_missed = []
    key_dict = {'FILE': 'Не выбран файл для подписания',
                'PFX': 'Не выбран сертификат для подписания',
                'PASSWORD': 'Не введён пароль для сертификата',
                'FOLDER': 'Не выбрана папка для сохранения подписанного документа',
                 }
    if values['type_file'] == 'EXCEL':
        key_dict['sheet_list'] = 'Не выбран лист для подписи'
    print(values)
    for key in key_dict:
        print(f'{key}  {values[key]}')
        if len(values[key]) < 1:

            keys_missed.append(key)

    if len(keys_missed) > 0:
        popup_string = 'Не удалось подписать потому что: \n'
        i = 0
        for key in keys_missed:
            i += 1
            popup_string += f'\n {i}. {key_dict[key]}'
        sg.PopupError(popup_string, title='Ахтунг!', keep_on_top=True)
        return False
    return True

def get_sheet_names(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    sheets = wb.sheetnames
    wb.close()
    return sheets
def excel_to_pdf(values):
    try:
        path_to_save = f'{values["FOLDER"]}/temp.pdf'
        excel = client.Dispatch("Excel.Application")
        excel.Interactive = False
        excel.Visible = False
        # excel.visible=0
        wb = excel.Workbooks.Open(values['FILE'])
        wb.application.displayalerts = False
        work_sheets = wb.Sheets(values['sheet_list'])
        # # Convert into PDF File
        try:
            work_sheets.ExportAsFixedFormat(0, path_to_save)
        except:
            os.remove(f'{values["FOLDER"]}/temp.pdf')
            time.sleep(0.1)
            work_sheets.ExportAsFixedFormat(0, path_to_save)
        wb.Close()
        excel.Quit()
        # window.write_event_value('-THREAD-',path_to_save)
        return path_to_save
    except Exception as e:
        print(e)
        excel.Quit()
        return False


if __name__ == '__main__':
    # import PySimpleGUI as sg
    #
    import win32api

    print(win32api.FormatMessage(-2147018887))
    # sg.theme_previewer()
    # subprocess.Popen([ r"C:/Users/skryabin.p/Desktop/СЗ_лист №2 (оборотный).pdf"],shell=True)